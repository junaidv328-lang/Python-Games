import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="FishyBiz · Fundamental Analyser v2",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── STYLE ─────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Instrument+Serif:ital@0;1&family=DM+Mono:wght@300;400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
:root {
    --ink:#0f0e0c; --paper:#f5f2ec; --accent:#c8401e; --accent2:#2a5c45;
    --muted:#7a7570; --border:#d6d0c8; --card:#ffffff;
    --pass:#2a5c45; --fail:#c8401e; --warn:#d4850a;
}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;background:var(--paper);color:var(--ink);}
h1,h2,h3{font-family:'Instrument Serif',serif;}
.main-header{font-family:'Instrument Serif',serif;font-size:2.6rem;font-style:italic;
  border-bottom:2px solid var(--ink);padding-bottom:0.4rem;margin-bottom:0.2rem;}
.sub-header{font-family:'DM Mono',monospace;font-size:0.7rem;letter-spacing:0.12em;
  color:var(--muted);text-transform:uppercase;margin-bottom:1.5rem;}
.stag{font-family:'DM Mono',monospace;font-size:0.62rem;letter-spacing:0.14em;
  text-transform:uppercase;color:var(--muted);background:var(--border);
  padding:0.18rem 0.5rem;border-radius:2px;display:inline-block;margin-bottom:0.4rem;}
.mcard{background:var(--card);border:1px solid var(--border);border-radius:4px;padding:1rem 1.2rem;}
.mval{font-family:'Instrument Serif',serif;font-size:1.7rem;font-style:italic;line-height:1.1;}
.mlbl{font-family:'DM Mono',monospace;font-size:0.62rem;letter-spacing:0.1em;
  text-transform:uppercase;color:var(--muted);margin-top:0.15rem;}
.rule-row{display:flex;align-items:center;padding:0.5rem 0;
  border-bottom:1px solid var(--border);gap:0.8rem;}
.rule-icon{font-size:1rem;width:1.3rem;flex-shrink:0;}
.rule-text{flex:1;font-size:0.84rem;}
.rule-val{font-family:'DM Mono',monospace;font-size:0.78rem;color:var(--muted);
  text-align:right;min-width:110px;}
.src-tag{font-family:'DM Mono',monospace;font-size:0.58rem;border:1px solid;
  padding:0 4px;border-radius:2px;white-space:nowrap;}
.flag-red{background:#fdf0ee;border-left:3px solid var(--fail);
  padding:0.7rem 1rem;border-radius:0 3px 3px 0;margin:0.3rem 0;font-size:0.83rem;}
.flag-green{background:#edf5f1;border-left:3px solid var(--pass);
  padding:0.7rem 1rem;border-radius:0 3px 3px 0;margin:0.3rem 0;font-size:0.83rem;}
.stButton>button{background:var(--ink);color:#fff;border:none;border-radius:3px;
  font-family:'DM Mono',monospace;font-size:0.78rem;letter-spacing:0.07em;padding:0.55rem 1.4rem;}
.stButton>button:hover{background:var(--accent);}
div[data-testid="stFileUploader"]{border:2px dashed var(--border);border-radius:4px;padding:1rem;}
hr{border-color:var(--border);}
</style>""", unsafe_allow_html=True)

# ── COLOUR PALETTE ────────────────────────────────────────────────────────────
PAPER="#f5f2ec"; INK="#0f0e0c"; A1="#c8401e"; A2="#2a5c45"; MUTED="#7a7570"
BLUE="#2b5f8e"; PURPLE="#6b3a8b"

SOURCE_COLORS = {
    "Graham":    "#1a4a8a",
    "Fisher":    "#2a5c45",
    "Siegel":    "#7a5c1a",
    "Dhandho":   "#8b1a1a",
    "Damodaran": "#1a5c5c",
}

# ── RULE EXPLANATIONS ─────────────────────────────────────────────────────────
RULE_EXPLANATIONS = {
    "No loss year": (
        "Profitability Consistency",
        "Net Profit > 0 in every reported year",
        "A company that has never reported a loss shows earnings stability. Graham insisted any loss year disqualifies a stock from 'investment' to 'speculation'. Loss years reveal hidden cyclicality, high fixed-cost leverage, or poor management."
    ),
    "PE ratio ≤ 20": (
        "P/E — Price-to-Earnings (Hard Cap)",
        "P/E = Market Price per Share ÷ Earnings per Share (EPS)",
        "P/E tells you how many rupees you pay for every ₹1 of annual profit. Graham set 20× as an absolute ceiling — above this you are paying a speculative premium. The higher the P/E, the more future growth must materialise to justify the price."
    ),
    "PE ratio ≤ 15": (
        "P/E — Fair Value Zone",
        "P/E = Market Price per Share ÷ EPS",
        "Graham's 'fair value' benchmark for an average business is 12–15×. Below 12× is cheap; 12–15× is fair; 15–20× is acceptable only for above-average businesses. This rule checks whether the stock is in the fair zone."
    ),
    "Price-to-Book": (
        "P/B — Price-to-Book Value",
        "P/B = Market Price per Share ÷ Book Value per Share (BVPS)",
        "Book Value is what remains for shareholders if the company paid off all debts and liquidated assets at accounting value. P/B < 1 means you buy below liquidation value — Graham called this 'net-net'. P/B ≤ 1.5× means a small justifiable premium over book. Above 3–4× requires strong earnings to justify."
    ),
    "Current Ratio": (
        "Current Ratio — Short-Term Liquidity",
        "Current Ratio = Current Assets ÷ Current Liabilities",
        "Current Assets are things that convert to cash within a year (cash, receivables, inventory). Current Liabilities are bills due within a year. Ratio ≥ 1.5 means ₹1.50 in liquid assets for every ₹1 owed short-term. Below 1.0 means the company may struggle to pay near-term obligations."
    ),
    "Debt-to-Equity": (
        "D/E — Debt-to-Equity Ratio",
        "D/E = Total Debt ÷ Shareholders' Equity",
        "Measures how much of the business is funded by borrowed money vs. owners' money. D/E < 1 means equity exceeds debt — conservatively financed. D/E > 2 means heavily leveraged; interest costs can wipe out profits during downturns. Graham preferred D/E below 0.5 for manufacturing companies."
    ),
    "Interest Coverage": (
        "ICR — Interest Coverage Ratio",
        "ICR = EBIT ÷ Interest Expense  (EBIT = Earnings Before Interest and Tax)",
        "ICR tells you how many times the company can pay its annual interest bill from operating profit. ICR of 3× means operating profit is 3× the interest due — comfortable buffer. ICR below 1.5× is dangerous: one bad year and the company cannot service its debt. Graham required ≥ 3× for industrial companies."
    ),
    "Consistent dividend": (
        "Dividend Track Record",
        "Dividend paid in ≥ 70% of reported years",
        "Graham treated dividend payment as a signal of financial discipline. Companies that consistently pay dividends cannot fake profitability — you need real cash to pay dividends. Irregular or zero dividends may mean the company retains cash for growth (acceptable if ROCE is high) or that profits are not real cash (concerning)."
    ),
    "OPM vs NPM gap": (
        "OPM vs NPM Gap — One-Timer Check",
        "Gap = |Operating Profit Margin − Net Profit Margin| in percentage points",
        "OPM (Operating Profit Margin) = Operating Profit ÷ Revenue. NPM (Net Profit Margin) = Net Profit ÷ Revenue. A large persistent gap suggests non-recurring items — asset sale gains, write-offs — are inflating or deflating reported profits. Graham insisted on stripping these out to see normalised earnings. Gap > 15pp triggers this flag."
    ),
    "Operating Margin ≥ 12%": (
        "OPM — Operating Profit Margin",
        "OPM = Operating Profit ÷ Revenue × 100",
        "OPM measures profit from core operations before interest and tax. Fisher (Point 5) said the greatest long-term gains come from companies with above-average and improving margins. 12% is the threshold for above-average across most industries. A company with 5% OPM has almost no buffer against cost increases or price cuts."
    ),
    "Operating margin trending": (
        "OPM Trend — Margin Expansion",
        "Slope of OPM over all reported years (linear regression)",
        "Fisher's Point 6: not just the level of margin but the direction matters. A company consistently expanding its operating margin shows pricing power, cost discipline, and management quality. Contracting margins — even from a high base — signal competitive pressure, input cost inflation, or inefficiency creeping in."
    ),
    "Net margin trending": (
        "NPM Trend — Net Profitability Direction",
        "Slope of NPM (Net Profit Margin) over all reported years",
        "Net Profit Margin = Net Profit ÷ Revenue. This captures the full bottom-line trend after interest and tax. Rising NPM confirms that operating improvement is not being eaten up by higher debt costs or tax. Falling NPM with rising OPM is a warning — leverage is increasing and eating the gains."
    ),
    "Revenue growth + margin": (
        "Operating Leverage — Quality Growth Signal",
        "Revenue growing (positive CAGR) AND OPM slope > 0, both simultaneously",
        "Operating leverage means: as revenue grows, margins improve because fixed costs are spread over more revenue. This is the gold standard of business quality — both volume growth and better unit economics simultaneously. Revenue growing but margins shrinking means the company is buying growth through discounting — not sustainable long-term."
    ),
    "Revenue CAGR": (
        "CAGR — Compound Annual Growth Rate of Revenue",
        "Revenue CAGR = (Revenue Last Year ÷ Revenue First Year) ^ (1 ÷ n) − 1, where n = number of years",
        "CAGR smooths out year-to-year volatility and shows the true underlying growth rate. Siegel's research shows long-run equity returns are anchored to earnings and revenue growth. 8% CAGR is the benchmark for a business growing faster than the broader economy (India nominal GDP ~6–7%). Below 8% suggests no real competitive advantage over the economy."
    ),
    "Earnings yield": (
        "Earnings Yield vs Risk-Free Rate (Siegel)",
        "Earnings Yield = (1 ÷ P/E) × 100. Compare to 10-yr Government Bond Yield.",
        "Earnings Yield is the inverse of P/E — the annual return you get from earnings at the current price. Siegel's insight: stocks should yield more than risk-free bonds to compensate for higher risk. If earnings yield < bond yield, you earn more from government bonds risk-free than from this stock. A basic sanity check on relative value."
    ),
    "Average ROCE": (
        "ROCE — Return on Capital Employed",
        "ROCE = EBIT ÷ Capital Employed × 100  (Capital Employed = Equity + Total Debt)",
        "ROCE measures how efficiently the company uses all its capital to generate operating profit. Pabrai (Dhandho) uses sustained high ROCE as the primary quantitative evidence of a competitive moat. A company earning 20%+ ROCE for 5+ years in a competitive market must have structural advantage — brand, patents, network effects, or switching costs. 15% is minimum; 20%+ signals a genuine moat."
    ),
    "FCF positive in": (
        "FCF — Free Cash Flow",
        "FCF = Cash from Operations (CFO) − Capital Expenditure (Capex)",
        "FCF is the real cash a business generates after maintaining and growing its asset base. It is money that can be returned to shareholders, used to reduce debt, or reinvested. Pabrai insists FCF must be positive and stable for the business to be truly cash-generative. Positive profits but negative FCF means the company is consuming cash faster than it earns it — often a working capital trap."
    ),
    "CFO ≥ 80% of Net Profit": (
        "CFO/NP — Cash Earnings Quality (Dhandho)",
        "Ratio = Cash from Operations ÷ Net Profit × 100",
        "CFO (Cash from Operations) is what actually hits the bank account. Net Profit is an accounting figure that includes non-cash items. When CFO is much lower than Net Profit, earnings are partly on paper — uncollected receivables, inventory build-up. Pabrai's rule: CFO should be ≥ 80% of Net Profit for earnings to be considered real. Below 60% consistently is a serious red flag."
    ),
    "Margin of Safety": (
        "MoS — Margin of Safety (Dhandho DCF)",
        "MoS = (Intrinsic Value − Market Cap) ÷ Intrinsic Value × 100. Intrinsic Value = 10-yr FCFF DCF at 10% growth + 12× terminal multiple.",
        "Pabrai's core principle from Dhandho: heads I win, tails I do not lose much. You only buy when market price is significantly below intrinsic value — the gap is your safety buffer against being wrong. 30% MoS means buying at 70 paise on the rupee. 50% MoS means 50 paise. The intrinsic value here is estimated using FCFF discounted at WACC with conservative assumptions."
    ),
    "EV/EBITDA": (
        "EV/EBITDA — Enterprise Value to EBITDA (Damodaran)",
        "EV = Market Cap + Total Debt − Cash. EBITDA = Operating Profit + Depreciation. EV/EBITDA = EV ÷ EBITDA.",
        "Damodaran considers EV/EBITDA superior to P/E because it is capital-structure neutral — it does not matter whether a company is funded by debt or equity. EV (Enterprise Value) is the total price to acquire the business including its debt. EBITDA represents raw operating cash generation before financing and accounting charges. ≤ 8× = cheap; 8–12× = fair; > 15× = expensive. Avoids distortions from different depreciation policies or debt levels."
    ),
    "PEG Ratio": (
        "PEG — Price/Earnings to Growth Ratio (Damodaran)",
        "PEG = P/E Ratio ÷ EPS Growth Rate (%). EPS CAGR computed from historical earnings per share.",
        "PEG adjusts the P/E ratio for earnings growth. A company with P/E of 30 growing earnings at 30% has PEG = 1.0 — you are paying a fair price for the growth you receive. PEG < 1 means the growth rate exceeds what you are paying for — potentially undervalued. PEG > 2 means growth is fully or over-priced. EPS = Earnings Per Share."
    ),
    "P/B justified by ROE": (
        "Fair P/B — Damodaran ROE-based Valuation Check",
        "Fair P/B = Average ROE ÷ Cost of Equity (%). Passes if actual P/B ≤ Fair P/B × 1.2 (20% tolerance).",
        "Damodaran's insight: a company is only worth more than its book value if it earns more than its cost of equity. ROE (Return on Equity) = Net Profit ÷ Shareholders' Equity. Cost of Equity = Risk-free rate + Beta × ERP. If ROE = 15% and CoE = 12%, fair P/B = 1.25×. Paying 3× book for a 15% ROE business is overpaying. ERP = Equity Risk Premium."
    ),
    "ROIC Spread": (
        "ROIC Spread — Value Creation Test (Damodaran Ch.11)",
        "ROIC Spread = ROCE − WACC (in percentage points). WACC = Weighted Average Cost of Capital.",
        "Damodaran's fundamental principle: a company only creates value when its return on invested capital exceeds its cost of capital. WACC is what debt-holders and equity-holders together require to earn. If ROCE > WACC: every rupee reinvested creates more than a rupee of value — growth is good. If ROCE < WACC: every rupee reinvested destroys value — the company would be worth more if it stopped growing and returned cash to shareholders."
    ),
    "EVA > 0": (
        "EVA — Economic Value Added (Damodaran Ch.32)",
        "EVA = NOPAT − (WACC × Capital Employed). NOPAT = Net Operating Profit After Tax = Operating Profit × (1 − Tax Rate).",
        "EVA is the rupee amount of value created above and beyond the cost of all capital used. A company can show accounting profit but negative EVA if it earns less than its cost of capital. Positive EVA means the company is genuinely enriching shareholders. Damodaran calls EVA the purest single-number summary of whether a business is worth owning. NOPAT = Net Operating Profit After Tax."
    ),
    "CFO / EBIT": (
        "CFO/EBIT — Damodaran Earnings Quality Check (Ch.3)",
        "Ratio = Cash from Operations ÷ EBIT (Operating Profit)",
        "EBIT (Earnings Before Interest and Tax) is the accounting operating profit. CFO is the actual cash generated. In a healthy business, CFO should equal or exceed EBIT because depreciation (a non-cash charge) is added back in CFO. CFO/EBIT < 1 means less cash than accounting profits suggest — a sign of aggressive revenue recognition, rising receivables, or inventory build-up. Damodaran flags this as accrual quality risk."
    ),
    "Working Capital / Revenue": (
        "WC Intensity — Working Capital Efficiency (Damodaran Ch.10)",
        "WC % of Revenue = (Current Assets − Current Liabilities) ÷ Revenue × 100. Trend should be flat or declining.",
        "Working Capital (WC) is cash tied up in day-to-day operations — inventory and receivables minus payables. As a % of revenue, rising WC means the company needs more cash just to grow the same amount. Damodaran calls this the 'implicit reinvestment trap' — businesses that need ever-more working capital consume cash invisibly. Declining WC/Revenue means the business gets more efficient as it scales — the ideal pattern."
    ),
    "Synthetic credit": (
        "Synthetic Credit Rating — Damodaran IC→Rating Map (Ch.8)",
        "ICR → Rating: >8.5×=AA, 6.5–8.5×=A, 4.5–6.5×=A−, 3.5–4.5×=BBB, 2.5–3.5×=BB, 1.5–2.5×=B, <1.5×=CCC",
        "Damodaran (Chapter 8) mapped interest coverage ratios to bond credit ratings for firms without public debt ratings. ICR (Interest Coverage Ratio) = Operating Profit ÷ Interest Expense. Investment grade (BBB and above) means the company comfortably services its debt and would qualify for institutional lending. Sub-investment grade (BB and below) means elevated default risk — higher cost of debt and potentially constrained credit access."
    ),
    "FCFF positive": (
        "FCFF — Free Cash Flow to Firm (Damodaran Ch.10, 15)",
        "FCFF = NOPAT + Depreciation − Capex − ΔWorking Capital. NOPAT = Operating Profit × (1 − Tax Rate). Δ = Change in.",
        "FCFF is the cash generated by the entire business before any payments to debt or equity holders. Damodaran uses FCFF as the basis for firm valuation (DCF). Unlike simple FCF (CFO − Capex), FCFF is tax-adjusted and accounts for working capital changes — making it more precise for intrinsic value estimation. A growing FCFF means the business's intrinsic value is compounding over time."
    ),
}


# ── CSV PARSER ────────────────────────────────────────────────────────────────
def parse_screener_csv(uploaded_file):
    """Parse Screener.in export — accepts both .xlsx and .csv files."""

    def sf(val):
        s = str(val).replace(",","").replace("%","").strip()
        if s in ("","-","—","N/A","na","NA","nan","#DIV/0!","none","None"): return np.nan
        try: return float(s)
        except: return np.nan

    filename = getattr(uploaded_file, "name", "")
    is_excel = filename.lower().endswith((".xlsx", ".xls"))

    parsed, company_name = {}, "Uploaded Company"

    if is_excel:
        # ── XLSX path — Screener.in Data Sheet format ──────────────────
        # Screener.in stores ALL data in "Data Sheet" tab in row-based sections.
        # Each section starts with a header row (e.g. 'PROFIT & LOSS') 
        # followed by a 'Report Date' row with year columns, then data rows.
        import openpyxl
        raw_bytes = uploaded_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)

        # ── Helper: format date cells to "Mar YYYY" ───────────────────
        import datetime as dt
        def fmt_date(v):
            if isinstance(v, (dt.datetime, dt.date)):
                return v.strftime("Mar %Y")
            return str(v).strip() if v is not None else ""

        # ── Try reading from "Data Sheet" (Screener.in v2 format) ──────
        data_sheet = None
        for sname in wb.sheetnames:
            if "data" in sname.lower():
                data_sheet = wb[sname]; break

        if data_sheet is not None:
            all_rows = list(data_sheet.iter_rows(values_only=True))

            # Extract company name
            for row in all_rows[:5]:
                if row[0] and "company" in str(row[0]).lower() and row[1]:
                    company_name = str(row[1]).strip(); break

            # Find section blocks: look for rows where col[0] is a section header
            # (all-caps like 'PROFIT & LOSS', 'BALANCE SHEET', 'CASH FLOW:')
            # followed immediately or soon by a 'Report Date' row
            section_markers = {
                "PROFIT & LOSS": "pl",
                "BALANCE SHEET": "bs",
                "CASH FLOW":     "cf",
            }

            def extract_section(all_rows, start_idx):
                """From start_idx, find the Report Date row and extract data as dict."""
                date_row_idx = None
                for i in range(start_idx, min(start_idx+10, len(all_rows))):
                    if all_rows[i][0] and "report date" in str(all_rows[i][0]).lower():
                        date_row_idx = i; break
                if date_row_idx is None:
                    return {}, []
                # Years from date row
                year_vals = [fmt_date(v) for v in all_rows[date_row_idx][1:] if v is not None]
                years_out = [y for y in year_vals if y]
                # Data rows until next empty or section header
                section_data = {}
                for i in range(date_row_idx+1, len(all_rows)):
                    row = all_rows[i]
                    if row[0] is None: continue
                    label = str(row[0]).strip()
                    if not label: continue
                    # Stop if we hit another section marker
                    label_up2 = label.upper().rstrip(":").strip()
                    stop_words = ["PROFIT & LOSS","BALANCE SHEET","CASH FLOW","QUARTERS","DERIVED","PRICE","META"]
                    if any(label_up2 == s or label_up2.startswith(s) for s in stop_words):
                        break
                    vals = list(row[1:])
                    numeric = {}
                    for j, y in enumerate(years_out):
                        v = vals[j] if j < len(vals) else None
                        numeric[y] = sf(v) if not isinstance(v, (dt.datetime, dt.date)) else np.nan
                    section_data[label] = numeric
                return section_data, years_out

            # Also extract META (current price, face value, shares)
            meta = {}
            for row in all_rows[:15]:
                if row[0] and row[1] is not None:
                    meta[str(row[0]).strip().lower()] = row[1]

            # Scan for section headers
            pl_data, bs_data, cf_data = {}, {}, {}
            for i, row in enumerate(all_rows):
                if row[0] is None: continue
                label_up = str(row[0]).strip().upper()
                if "PROFIT" in label_up and "LOSS" in label_up:
                    pl_data, _ = extract_section(all_rows, i)
                elif "BALANCE" in label_up:
                    bs_data, _ = extract_section(all_rows, i)
                elif "CASH" in label_up and ("FLOW" in label_up or label_up.startswith("CASH")):
                    cf_data, _ = extract_section(all_rows, i)

            # Build parsed sections as DataFrames for compatibility
            def dict_to_df(section_dict, section_name):
                if not section_dict: return pd.DataFrame()
                years_list = list(next(iter(section_dict.values())).keys())
                rows = []
                for label, yr_vals in section_dict.items():
                    row = [label] + [yr_vals.get(y, np.nan) for y in years_list]
                    rows.append(row)
                cols = [section_name] + years_list
                return pd.DataFrame(rows, columns=cols)

            # Compute derived rows from Data Sheet
            # 1. Total Equity = Equity Share Capital + Reserves
            if bs_data:
                sc_row = bs_data.get("Equity Share Capital", {})
                res_row = bs_data.get("Reserves", {})
                if sc_row and res_row:
                    all_yrs_bs = list(sc_row.keys())
                    equity_row = {}
                    for y in all_yrs_bs:
                        sc_v = sc_row.get(y, np.nan)
                        re_v = res_row.get(y, np.nan)
                        equity_row[y] = (sc_v + re_v) if not np.isnan(sc_v) and not np.isnan(re_v) else re_v
                    bs_data["Total Equity"] = equity_row

            # 1b. EPS and PE from shares and net profit
            # Data Sheet stores "No. of Equity Shares" and "PRICE:" rows
            shares_row = {}
            price_hist_row = {}
            adj_shares_cr = {}
            for i2, row2 in enumerate(all_rows):
                if row2[0] is None: continue
                lbl2 = str(row2[0]).strip()
                if "no. of equity shares" in lbl2.lower() or "number of equity shares" in lbl2.lower():
                    # pair with the annual years
                    yrs_tmp = list(pl_data.get("Net profit", {}).keys()) if pl_data else []
                    for k2, y2 in enumerate(yrs_tmp):
                        v2 = row2[k2+1] if k2+1 < len(row2) else None
                        shares_row[y2] = sf(v2)
                if lbl2.upper().startswith("PRICE"):
                    yrs_tmp = list(pl_data.get("Net profit", {}).keys()) if pl_data else []
                    for k2, y2 in enumerate(yrs_tmp):
                        v2 = row2[k2+1] if k2+1 < len(row2) else None
                        price_hist_row[y2] = sf(v2)
                if "adjusted equity shares" in lbl2.lower():
                    yrs_tmp = list(pl_data.get("Net profit", {}).keys()) if pl_data else []
                    for k2, y2 in enumerate(yrs_tmp):
                        v2 = row2[k2+1] if k2+1 < len(row2) else None
                        adj_shares_cr[y2] = sf(v2)  # already in Cr

            if pl_data and (shares_row or adj_shares_cr):
                np_row2 = pl_data.get("Net profit", {})
                eps_row = {}
                pe_row  = {}
                pb_row  = {}
                for y2 in np_row2:
                    np_v2 = np_row2.get(y2, np.nan)
                    # Use adjusted shares (in Cr) if available, else convert raw shares
                    if adj_shares_cr.get(y2) and not np.isnan(adj_shares_cr.get(y2, np.nan)):
                        shares_cr = adj_shares_cr[y2]  # already in crores
                    elif shares_row.get(y2) and not np.isnan(shares_row.get(y2, np.nan)):
                        shares_cr = shares_row[y2] / 1e7  # convert to crores
                    else:
                        shares_cr = np.nan
                    # EPS = Net Profit (Cr) / Shares (Cr) = ₹ per share
                    if not np.isnan(np_v2) and not np.isnan(shares_cr) and shares_cr > 0:
                        eps_v = np_v2 / shares_cr
                        eps_row[y2] = eps_v
                        # PE = Historical Price / EPS
                        hist_price = price_hist_row.get(y2, np.nan)
                        # PRICE: row stores price in hundreds (e.g., 2.42 = ₹242?)
                        # Check: Cupid current price = ₹112.5, PRICE: last value = 12.53
                        # Ratio = 112.5/12.53 ≈ 8.98... seems like multiplier needed
                        # Actually the PRICE: row likely stores price / face_value or price directly
                        # Let's check: market cap = 15127.45 Cr, shares = 268.47 Cr
                        # Price = 15127.45 / 268.47 = ₹56.35... but current price = 112.5
                        # PRICE: last value = 12.53 → × face value 1 = 12.53? No
                        # More likely PRICE: = Price/EPS = PE ratio stored directly!
                        # Check: PE = 12.53, EPS = 112.5/12.53 = 8.98... let's verify
                        # Net profit 2025 = 40.93 Cr, adj shares = 134.23 Cr
                        # EPS = 40.93/134.23 = 0.305 ₹... × 100 = 30.5? 
                        # Wait face value changed — 2024 onwards face=1, before face=10
                        # So current EPS = 40.93Cr / 268.47Cr shares = ₹1.525
                        # PE = 112.5 / 1.525 = 73.8x... PRICE: shows 12.53
                        # Conclusion: PRICE: row IS the PE ratio stored by Screener!
                        if not np.isnan(hist_price) and hist_price > 0:
                            pe_row[y2] = hist_price  # PRICE: row = PE ratio directly
                    else:
                        eps_row[y2] = np.nan

                if eps_row:
                    pl_data["EPS"] = eps_row
                if pe_row:
                    pl_data["PE Ratio"] = pe_row

            # 2. Operating Profit = Profit before tax + Interest - Other Income
            if pl_data:
                pbt = pl_data.get("Profit before tax", {})
                intr = pl_data.get("Interest", {})
                oinc = pl_data.get("Other Income", {})
                dep  = pl_data.get("Depreciation", {})
                if pbt:
                    yrs_pl = list(pbt.keys())
                    op_row = {}
                    for y in yrs_pl:
                        pbt_v = pbt.get(y, np.nan)
                        int_v = intr.get(y, 0) if intr else 0
                        oi_v  = oinc.get(y, 0) if oinc else 0
                        dep_v = dep.get(y, 0) if dep else 0
                        if not np.isnan(pbt_v):
                            # Op Profit = PBT + Interest + Depreciation - Other Income
                            op_row[y] = pbt_v + (int_v if not np.isnan(int_v) else 0) + (dep_v if not np.isnan(dep_v) else 0) - (oi_v if not np.isnan(oi_v) else 0)
                        else:
                            op_row[y] = np.nan
                    pl_data["Operating Profit"] = op_row
                # OPM = Operating Profit / Sales
                sales_row = pl_data.get("Sales", {})
                if sales_row and "Operating Profit" in pl_data:
                    opm_row = {}
                    for y in yrs_pl:
                        op_v = pl_data["Operating Profit"].get(y, np.nan)
                        s_v  = sales_row.get(y, np.nan)
                        opm_row[y] = (op_v/s_v*100) if not np.isnan(op_v) and not np.isnan(s_v) and s_v>0 else np.nan
                    pl_data["OPM"] = opm_row
                # NPM = Net Profit / Sales
                np_row = pl_data.get("Net profit", {})
                if np_row and sales_row:
                    npm_row = {}
                    for y in yrs_pl:
                        np_v = np_row.get(y, np.nan)
                        s_v  = sales_row.get(y, np.nan)
                        npm_row[y] = (np_v/s_v*100) if not np.isnan(np_v) and not np.isnan(s_v) and s_v>0 else np.nan
                    pl_data["NPM"] = npm_row
                # EPS from net profit and shares (if available from meta)
                # Interest coverage = Operating Profit / Interest
                if intr and "Operating Profit" in pl_data:
                    icr_row = {}
                    for y in yrs_pl:
                        op_v = pl_data["Operating Profit"].get(y, np.nan)
                        i_v  = intr.get(y, np.nan)
                        icr_row[y] = (op_v/i_v) if not np.isnan(op_v) and not np.isnan(i_v) and i_v>0 else np.nan
                    pl_data["Interest Coverage"] = icr_row

            # 3. ROCE = Operating Profit / (Equity + Debt) — compute and store in bs_data
            if pl_data and bs_data and "Operating Profit" in pl_data:
                eq_r  = bs_data.get("Total Equity", {})
                td_r  = bs_data.get("Borrowings", {})
                op_r  = pl_data.get("Operating Profit", {})
                yrs_  = list(op_r.keys())
                roce_row = {}
                roe_row  = {}
                np_r  = pl_data.get("Net profit", {})
                for y in yrs_:
                    eq_v = eq_r.get(y, np.nan)
                    td_v = td_r.get(y, 0) if td_r else 0
                    if np.isnan(td_v): td_v = 0
                    op_v = op_r.get(y, np.nan)
                    np_v = np_r.get(y, np.nan) if np_r else np.nan
                    ce = eq_v + td_v if not np.isnan(eq_v) else np.nan
                    roce_row[y] = (op_v/ce*100) if not np.isnan(op_v) and not np.isnan(ce) and ce>0 else np.nan
                    roe_row[y]  = (np_v/eq_v*100) if not np.isnan(np_v) and not np.isnan(eq_v) and eq_v>0 else np.nan
                bs_data["ROCE"] = roce_row
                bs_data["ROE"]  = roe_row

            # Direct cash flow extraction — scan for "Cash from Operating Activity" anywhere
            if not cf_data:
                cf_years = None
                for i, row in enumerate(all_rows):
                    if row[0] and "report date" in str(row[0]).lower() and cf_years is None:
                        # Only use the date row that comes after a "CASH" header
                        for j in range(max(0,i-5), i):
                            if all_rows[j][0] and "cash" in str(all_rows[j][0]).lower():
                                cf_years = [fmt_date(v) for v in row[1:] if v is not None]
                                break
                    if row[0] and "cash from operating" in str(row[0]).lower():
                        if cf_years:
                            vals = list(row[1:])
                            cf_data["Cash from Operating Activity"] = {
                                y: sf(vals[k]) if k < len(vals) else np.nan
                                for k,y in enumerate(cf_years)
                            }
                    if row[0] and "cash from investing" in str(row[0]).lower():
                        if cf_years:
                            vals = list(row[1:])
                            cf_data["Cash from Investing Activity"] = {
                                y: sf(vals[k]) if k < len(vals) else np.nan
                                for k,y in enumerate(cf_years)
                            }

            if pl_data: parsed["Profit & Loss"] = dict_to_df(pl_data, "Profit & Loss")
            if bs_data: parsed["Balance Sheet"] = dict_to_df(bs_data, "Balance Sheet")
            if cf_data: parsed["Cash Flow"] = dict_to_df(cf_data, "Cash Flow")

            # Store meta for later use (current price etc.)
            parsed["__meta__"] = meta

        else:
            # ── Fallback: try reading individual sheets ────────────────
            def sheet_to_df(ws, sec_name):
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2: return pd.DataFrame()
                # Find a row that looks like a date header
                header_idx = None
                for i, row in enumerate(rows):
                    non_null = [v for v in row[1:] if v is not None]
                    if non_null and isinstance(non_null[0], (dt.datetime, dt.date)):
                        header_idx = i; break
                if header_idx is None: return pd.DataFrame()
                years = [fmt_date(v) for v in rows[header_idx][1:] if v is not None]
                data_rows = []
                for row in rows[header_idx+1:]:
                    if row[0] is None: continue
                    vals = [str(row[0]).strip()] + [sf(v) for v in row[1:len(years)+1]]
                    data_rows.append(vals)
                if not data_rows: return pd.DataFrame()
                return pd.DataFrame(data_rows, columns=[sec_name]+years)

            for sname in wb.sheetnames:
                ws = wb[sname]
                sname_lo = sname.lower()
                if "profit" in sname_lo:
                    df = sheet_to_df(ws, sname)
                    if not df.empty: parsed[sname] = df
                elif "balance" in sname_lo:
                    df = sheet_to_df(ws, sname)
                    if not df.empty: parsed[sname] = df
                elif "cash" in sname_lo:
                    df = sheet_to_df(ws, sname)
                    if not df.empty: parsed[sname] = df

    else:
        # ── CSV path (original logic) ──────────────────────────────────
        content = uploaded_file.read()
        try:    text = content.decode("utf-8")
        except: text = content.decode("latin-1")

        lines = text.splitlines()
        raw_sections, cur = [], []
        for line in lines:
            if line.strip() == "":
                if cur: raw_sections.append(cur); cur=[]
            else: cur.append(line)
        if cur: raw_sections.append(cur)

        for block in raw_sections:
            try:
                df = pd.read_csv(io.StringIO("\n".join(block)))
                if df.empty or len(df.columns) < 2: continue
                parsed[str(df.columns[0]).strip()] = df
            except: continue

    for name in parsed:
        if name.startswith("__"): continue
        if not any(kw in name.lower() for kw in ["profit","balance","cash","ratio","quarter","annual"]):
            company_name = name; break
    # For Screener.in xlsx, company name comes from meta
    if company_name == "Uploaded Company" and "__meta__" in parsed:
        meta = parsed["__meta__"]
        for k,v in meta.items():
            if "company" in k.lower() and v:
                company_name = str(v).strip(); break

    def find(keys):
        for k in parsed:
            for key in keys:
                if key.lower() in k.lower(): return parsed[k]
        return pd.DataFrame()

    def to_series(df):
        res = {}
        if df.empty or len(df.columns)<2: return res
        lc = df.columns[0]
        for _,row in df.iterrows():
            label = str(row[lc]).strip()
            res[label] = {str(yc).strip(): sf(row[yc]) for yc in df.columns[1:]}
        return res

    pl  = to_series(find(["profit & loss","profit and loss","income statement","annual p&l"]))
    bs  = to_series(find(["balance sheet","balance"]))
    cf  = to_series(find(["cash flow","cashflow"]))
    rat = to_series(find(["ratio","ratios"]))

    years = list(pl[next(iter(pl))].keys()) if pl else []

    def gm(d, *keys):
        for k in d:
            for pk in keys:
                if pk.lower() in k.lower(): return d[k]
        return {}

    # For Screener.in Data Sheet format, some ratios live in balance sheet section
    def gm_multi(d1, d2, d3, *keys):
        """Try multiple dicts for a metric."""
        for d in [d1, d2, d3]:
            r = gm(d, *keys)
            if r: return r
        return {}

    rev   = gm(pl,"sales","revenue","net sales","net revenue")
    np_   = gm(pl,"net profit","pat","profit after tax","net income","profit for")
    op_   = gm(pl,"operating profit","ebit","pbdit","ebitda")
    intr  = gm(pl,"interest","finance cost","finance charges","interest expense")
    dep   = gm(pl,"depreciation","amortis","depreciation & amortis")
    tax_  = gm(pl,"tax","income tax","provision for tax")
    div_  = gm(pl,"dividend","equity dividend")
    eps_  = gm(pl,"eps","earning per share","earnings per share","diluted eps")

    ta_   = gm(bs,"total asset","net asset","total assets")
    eq_   = gm_multi(bs,rat,pl,"equity share capital","shareholder","net worth","shareholders fund","total equity","reserves")
    td_   = gm(bs,"total debt","borrowings","long term borrowing","total borrowing","borrowing")
    ca_   = gm(bs,"current assets","current asset","total current asset")
    cl_   = gm(bs,"current liab","current liability","total current liab")
    res_  = gm(bs,"reserve","reserves and surplus","other equity")
    sc_   = gm(bs,"share capital","paid up capital","equity share capital","equity share")
    fa_   = gm(bs,"fixed assets","net block","property plant","tangible assets")
    cash_ = gm(bs,"cash and cash","cash equivalent","cash & bank","cash balance","cash & bank")
    rec_  = gm(bs,"receivable","debtors","trade receivable","receivables")
    inv_  = gm(bs,"inventor","stock in trade","inventories")

    cfo_  = gm_multi(cf,pl,bs,"cash from oper","operating activities","net cash from oper","cash from operating activity","operating activity")
    cfi_  = gm(cf,"cash from invest","investing activities")
    cpx_  = gm_multi(cf,pl,bs,"capex","capital expenditure","purchase of fixed","cash from investing activity","investing activity","cash from investing")

    pe_   = gm_multi(rat,pl,bs,"p/e","pe ratio","price to earning","price earning")
    bv_   = gm(rat,"book value","bvps","book val per share")
    pb_   = gm_multi(rat,pl,bs,"p/b","pb ratio","price to book","price/book")
    roce_ = gm_multi(bs,rat,pl,"roce","return on capital employed","return on capital emp","return on capital")
    roe_  = gm_multi(bs,rat,pl,"roe","return on equity","return on net worth","return on networth")
    de_   = gm_multi(rat,bs,pl,"debt to equity","d/e","debt/equity","debt-to-equity")
    cr_   = gm_multi(rat,bs,pl,"current ratio")
    opm_  = gm_multi(rat,pl,bs,"opm","operating profit margin","operating margin","ebit margin")
    npm_  = gm_multi(rat,pl,bs,"npm","net profit margin","net margin","pat margin")

    # Extract current price from Screener meta (stored in Data Sheet row 8)
    screener_price = np.nan
    if "__meta__" in parsed:
        for k, v in parsed["__meta__"].items():
            if "current price" in k.lower() or "price" in k.lower():
                try: screener_price = float(v)
                except: pass
                break

    return dict(company_name=company_name, years=years, screener_price=screener_price,
        revenue=rev, net_profit=np_, operating_profit=op_,
        interest=intr, depreciation=dep, tax=tax_, dividend=div_, eps=eps_,
        total_assets=ta_, equity=eq_, total_debt=td_,
        current_assets=ca_, current_liab=cl_, reserves=res_,
        share_capital=sc_, fixed_assets=fa_, cash_and_equiv=cash_,
        receivables=rec_, inventory=inv_,
        cfo=cfo_, cfi=cfi_, capex=cpx_,
        pe_ratio=pe_, book_val=bv_, pb_ratio=pb_,
        roce_ratio=roce_, roe_ratio=roe_, debt_eq_r=de_,
        current_r=cr_, opm=opm_, npm=npm_)

# ── HELPERS ───────────────────────────────────────────────────────────────────
def sv(d, y): return d.get(y, np.nan) if d else np.nan
def av(d, ys): return [sv(d,y) for y in ys]
def smean(lst):
    c=[x for x in lst if not np.isnan(x)]; return np.mean(c) if c else np.nan
def sdiv(a,b):
    if any(np.isnan(x) for x in [a,b]) or b==0: return np.nan
    return a/b
def slope(vals):
    c=[(i,v) for i,v in enumerate(vals) if not np.isnan(v)]
    if len(c)<2: return np.nan
    xs=np.array([x[0] for x in c],float); ys=np.array([x[1] for x in c],float)
    return np.polyfit(xs,ys,1)[0]
def latest(lst): return next((v for v in reversed(lst) if not np.isnan(v)), np.nan)
def nn(v): return np.isnan(v)

# ── SYNTHETIC CREDIT RATING ───────────────────────────────────────────────────
def synthetic_rating(ic):
    if nn(ic) or ic==999: return "AAA", 0.59
    if ic>8.5:  return "AA",  0.70
    if ic>6.5:  return "A",   1.07
    if ic>4.5:  return "A−",  1.21
    if ic>3.5:  return "BBB", 1.47
    if ic>2.5:  return "BB",  2.21
    if ic>1.5:  return "B",   3.61
    return "CCC", 8.51

# ── ANALYSIS ENGINE ───────────────────────────────────────────────────────────
def run_analysis(data, current_price=None, bond_yield=7.0, erp=5.5, beta=1.0):
    years = data["years"]
    result = {"rules":[], "flags":[], "positives":[], "score":0, "max_score":0}

    def add(name, source, passed, pts, actual, threshold, note="", w=1):
        p = pts*w if passed else 0
        result["rules"].append(dict(name=name,source=source,passed=passed,
            pts=p,max_pts=pts*w,actual=actual,threshold=threshold,note=note))
        result["score"]     += p
        result["max_score"] += pts*w
        if note:
            (result["positives"] if passed else result["flags"]).append(f"[{source}] {name}: {note}")

    # ── raw series ────────────────────────────────────────────────────
    np_v  = av(data["net_profit"],   years)
    op_v  = av(data["operating_profit"], years)
    rev_v = av(data["revenue"],      years)
    int_v = av(data["interest"],     years)
    dep_v = av(data["depreciation"], years)
    tax_v = av(data["tax"],          years)
    div_v = av(data["dividend"],     years)
    eps_v = av(data["eps"],          years)
    ca_v  = av(data["current_assets"], years)
    cl_v  = av(data["current_liab"],   years)
    td_v  = av(data["total_debt"],     years)
    eq_v  = av(data["equity"],         years)
    ta_v  = av(data["total_assets"],   years)
    sc_v  = av(data["share_capital"],  years)
    fa_v  = av(data["fixed_assets"],   years)
    cash_v= av(data["cash_and_equiv"], years)
    rec_v = av(data["receivables"],    years)
    inv_v = av(data["inventory"],      years)
    cfo_v = av(data["cfo"],    years)
    cpx_v = av(data["capex"],  years)
    pe_v  = av(data["pe_ratio"],  years)
    pb_v  = av(data["pb_ratio"],  years)
    roce_v= av(data["roce_ratio"],years)
    roe_v = av(data["roe_ratio"], years)
    de_v  = av(data["debt_eq_r"], years)
    cr_v  = av(data["current_r"], years)
    opm_v = av(data["opm"],  years)
    npm_v = av(data["npm"],  years)

    # ── derived series ────────────────────────────────────────────────
    # FCF = CFO - |Capex|
    fcf_v = [c-abs(cx) if not nn(c) and not nn(cx) else (c if not nn(c) else np.nan)
             for c,cx in zip(cfo_v, cpx_v)]
    # EBITDA = Operating Profit + Depreciation
    ebitda_v = [o+d if not nn(o) and not nn(d) else np.nan for o,d in zip(op_v,dep_v)]
    # Working capital
    wc_v = [c-l if not nn(c) and not nn(l) else np.nan for c,l in zip(ca_v,cl_v)]
    # WC % of Revenue
    wc_pct_v = [sdiv(w,r)*100 for w,r in zip(wc_v,rev_v)]
    # Tax rate
    tax_rate_v = [sdiv(t, o+t) if not nn(t) and not nn(o) else np.nan for t,o in zip(tax_v, np_v)]
    avg_tax = smean([v for v in tax_rate_v if not nn(v) and 0<v<1])
    if nn(avg_tax): avg_tax = 0.25
    # NOPAT = Operating Profit * (1 - tax rate)
    nopat_v = [o*(1-avg_tax) if not nn(o) else np.nan for o in op_v]
    # Capital Employed = Total Assets - Current Liabilities (or Equity + Debt)
    ce_v = [e+d if not nn(e) and not nn(d) else np.nan for e,d in zip(eq_v,td_v)]
    # FCFF = NOPAT + Dep - Capex - delta WC
    fcff_v = []
    for i,y in enumerate(years):
        no = nopat_v[i]; d = dep_v[i]; cx = cpx_v[i]
        dwc = wc_v[i]-wc_v[i-1] if i>0 and not nn(wc_v[i]) and not nn(wc_v[i-1]) else np.nan
        if not nn(no) and not nn(d) and not nn(cx) and not nn(dwc):
            fcff_v.append(no + d - abs(cx) - dwc)
        elif not nn(no) and not nn(d) and not nn(cx):
            fcff_v.append(no + d - abs(cx))
        else:
            fcff_v.append(np.nan)

    # latest values
    l_np   = latest(np_v)
    l_op   = latest(op_v)
    l_rev  = latest(rev_v)
    l_int  = latest(int_v)
    l_dep  = latest(dep_v)
    l_ca   = latest(ca_v)
    l_cl   = latest(cl_v)
    l_td   = latest(td_v)
    l_eq   = latest(eq_v)
    l_ta   = latest(ta_v)
    l_sc   = latest(sc_v)
    l_cash = latest(cash_v)
    l_cfo  = latest(cfo_v)
    l_pe   = latest(pe_v)
    l_pb   = latest(pb_v)
    l_roce = latest(roce_v)
    l_roe  = latest(roe_v)
    l_de   = latest(de_v)
    l_cr   = latest(cr_v)
    l_opm  = latest(opm_v)
    l_npm  = latest(npm_v)
    l_ebitda = latest(ebitda_v)
    l_ce   = latest(ce_v)
    l_nopat= latest(nopat_v)
    l_fcff = latest(fcff_v)

    # Derive current ratio if missing
    if nn(l_cr) and not nn(l_ca) and not nn(l_cl) and l_cl>0:
        l_cr = l_ca / l_cl
    # Derive D/E if missing
    if nn(l_de) and not nn(l_td) and not nn(l_eq) and l_eq>0:
        l_de = l_td / l_eq
    # Interest coverage
    ic = sdiv(l_op, l_int)
    if not nn(l_int) and l_int < 1: ic = 999
    # After-tax cost of debt
    atcd = sdiv(l_int, l_td) * (1 - avg_tax) if not nn(l_int) and not nn(l_td) and l_td>0 else 0.06*(1-avg_tax)
    # Estimated WACC
    if not nn(l_eq) and not nn(l_td) and (l_eq+l_td)>0:
        we = l_eq/(l_eq+l_td); wd = l_td/(l_eq+l_td)
    else:
        we, wd = 0.7, 0.3
    cost_of_equity = (bond_yield/100) + beta * (erp/100)
    wacc = cost_of_equity*we + atcd*wd
    wacc_pct = wacc*100
    # ROIC spread
    roic_spread = (l_roce - wacc_pct) if not nn(l_roce) else np.nan
    # EVA
    eva = (l_nopat - wacc*l_ce) if not nn(l_nopat) and not nn(l_ce) else np.nan

    # ════════════════════════════════════════════════════════════════════
    # SECTION A — GRAHAM (safety & value)
    # ════════════════════════════════════════════════════════════════════
    loss_yrs = [years[i] for i,v in enumerate(np_v) if not nn(v) and v<0]
    add("No loss year in entire history","Graham",len(loss_yrs)==0, 8,
        "None" if not loss_yrs else f"{len(loss_yrs)} loss year(s)",
        "0 loss years",
        note="Consistent profitability" if not loss_yrs else f"Loss in: {', '.join(loss_yrs)}")

    pe_ok = not nn(l_pe) and l_pe<=20
    add("PE ratio ≤ 20× (hard ceiling)","Graham",pe_ok, 7,
        f"{l_pe:.1f}×" if not nn(l_pe) else "N/A","≤ 20×",
        note="Within Graham's investment cap" if pe_ok else "Above Graham's hard ceiling — speculative price")

    pe_fair = not nn(l_pe) and l_pe<=15
    add("PE ratio ≤ 15× (fair value zone)","Graham",pe_fair, 4,
        f"{l_pe:.1f}×" if not nn(l_pe) else "N/A","≤ 15×",
        note="In Graham's 'typical fair' range (12–15×)" if pe_fair else "Above fair-value zone")

    pb_ok = not nn(l_pb) and l_pb<=1.5
    add("Price-to-Book ≤ 1.5×","Graham",pb_ok, 5,
        f"{l_pb:.2f}×" if not nn(l_pb) else "N/A","≤ 1.5×",
        note="Deep value at P/B < 1 (net-net)" if not nn(l_pb) and l_pb<1 else ("Asset-backed pricing" if pb_ok else "Price well above book value"))

    cr_ok = not nn(l_cr) and l_cr>=1.5
    add("Current Ratio ≥ 1.5","Graham",cr_ok, 5,
        f"{l_cr:.2f}" if not nn(l_cr) else "N/A","≥ 1.5",
        note="Liquid balance sheet" if cr_ok else "Liquidity is tight")

    de_ok = not nn(l_de) and l_de<1.0
    add("Debt-to-Equity < 1.0","Graham",de_ok, 6,
        f"{l_de:.2f}" if not nn(l_de) else "N/A","< 1.0",
        note="Conservative leverage" if de_ok else "Leverage is elevated")

    ic_ok = not nn(ic) and ic>=3.0
    add("Interest Coverage ≥ 3×","Graham",ic_ok, 5,
        f"{ic:.1f}×" if not nn(ic) and ic!=999 else ("Debt-free" if ic==999 else "N/A"),"≥ 3×",
        note="Fixed charges comfortably covered" if ic_ok else "Thin interest cover — debt risk")

    dpaid = sum(1 for v in div_v if not nn(v) and v>0)
    div_ok = dpaid/len(years)>=0.7 if years else False
    add("Consistent dividend (≥70% of years)","Graham",div_ok, 4,
        f"{dpaid}/{len(years)} years","≥70% of years",
        note="Reliable shareholder returns" if div_ok else "Irregular / no dividend track record")

    opm_avg = smean([v for v in opm_v if not nn(v)])
    npm_avg = smean([v for v in npm_v if not nn(v)])
    gap = abs(opm_avg-npm_avg) if not nn(opm_avg) and not nn(npm_avg) else np.nan
    gap_ok = not nn(gap) and gap<15
    add("OPM vs NPM gap < 15pp (one-timer check)","Graham",gap_ok,3,
        f"{gap:.1f}pp" if not nn(gap) else "N/A","< 15pp",
        note="Earnings appear clean" if gap_ok else "Large gap — check for non-recurring items")

    # ════════════════════════════════════════════════════════════════════
    # SECTION B — FISHER (quality & growth)
    # ════════════════════════════════════════════════════════════════════
    opm_ok = not nn(l_opm) and l_opm>=12
    add("Operating Margin ≥ 12%","Fisher",opm_ok, 5,
        f"{l_opm:.1f}%" if not nn(l_opm) else "N/A","≥ 12%",
        note="Above-average margin — competitive strength" if opm_ok else "Thin margins — marginal business")

    sl_opm = slope([v for v in opm_v if not nn(v)])
    opm_trend_ok = not nn(sl_opm) and sl_opm>=0
    add("Operating margin trending upward","Fisher",opm_trend_ok, 4,
        f"{'Expanding' if opm_trend_ok else 'Contracting'} ({sl_opm:+.2f}pp/yr)" if not nn(sl_opm) else "N/A",
        "Positive slope",
        note="Management improving profitability over time" if opm_trend_ok else "Margins under pressure")

    sl_npm = slope([v for v in npm_v if not nn(v)])
    npm_trend_ok = not nn(sl_npm) and sl_npm>=0
    add("Net margin trending upward","Fisher",npm_trend_ok, 3,
        f"{'Expanding' if npm_trend_ok else 'Contracting'} ({sl_npm:+.2f}pp/yr)" if not nn(sl_npm) else "N/A",
        "Positive slope",
        note="Sustainable net profitability improvement" if npm_trend_ok else "Net margins declining")

    # Revenue + Margin expansion = positive operating leverage (Fisher/Damodaran cross)
    clean_rev = [v for v in rev_v if not nn(v)]
    rev_growing = len(clean_rev)>=2 and clean_rev[-1]>clean_rev[0]
    op_lev_ok = rev_growing and opm_trend_ok
    add("Revenue growth + margin expansion (operating leverage)","Fisher",op_lev_ok, 4,
        "Positive" if op_lev_ok else ("Rev up, margin down" if rev_growing else "Revenue not growing"),
        "Both expanding",
        note="Sales growth with improving margins — pricing power confirmed" if op_lev_ok else "Revenue may be growing but not translating to better margins")

    # ════════════════════════════════════════════════════════════════════
    # SECTION C — SIEGEL (market valuation)
    # ════════════════════════════════════════════════════════════════════
    rev_cagr = np.nan
    if len(clean_rev)>=3 and clean_rev[0]>0:
        rev_cagr = (clean_rev[-1]/clean_rev[0])**(1/(len(clean_rev)-1))-1
    cagr_ok = not nn(rev_cagr) and rev_cagr>=0.08
    add("Revenue CAGR ≥ 8%","Siegel",cagr_ok, 4,
        f"{rev_cagr*100:.1f}%" if not nn(rev_cagr) else "N/A","≥ 8% CAGR",
        note=f"Solid top-line growth ({rev_cagr*100:.1f}%)" if cagr_ok else "Revenue growth below benchmark")

    if not nn(l_pe) and l_pe>0:
        ey = (1/l_pe)*100
        ey_ok = ey>bond_yield
        add(f"Earnings yield > Bond yield ({bond_yield:.1f}%)","Siegel",ey_ok, 4,
            f"EY={ey:.1f}%","f> {bond_yield:.1f}%",
            note="Stock yields more than risk-free bond" if ey_ok else "Bond more attractive at this price")

    # ════════════════════════════════════════════════════════════════════
    # SECTION D — DHANDHO (moat & cash)
    # ════════════════════════════════════════════════════════════════════
    avg_roce = smean([v for v in roce_v if not nn(v)])
    roce_moat = not nn(avg_roce) and avg_roce>=20
    roce_ok   = not nn(avg_roce) and avg_roce>=15
    add("Average ROCE ≥ 15% (moat proxy)","Dhandho",roce_ok, 7,
        f"{avg_roce:.1f}%" if not nn(avg_roce) else "N/A","≥ 15%",
        note=("Strong moat: avg ROCE ≥ 20%" if roce_moat else "Decent capital returns") if roce_ok else "Low ROCE — no clear moat")

    fcf_pos = sum(1 for v in fcf_v if not nn(v) and v>0)
    fcf_tot  = sum(1 for v in fcf_v if not nn(v))
    fcf_ok   = fcf_tot>0 and fcf_pos/fcf_tot>=0.7
    add("FCF positive in ≥70% of years","Dhandho",fcf_ok, 6,
        f"{fcf_pos}/{fcf_tot} years positive","≥70% positive",
        note="Business generates real cash consistently" if fcf_ok else "FCF erratic — working capital or capex issues")

    cfo_np_ratio = sdiv(l_cfo, l_np)*100 if not nn(l_np) and l_np!=0 else np.nan
    cfo_np_ok = not nn(cfo_np_ratio) and cfo_np_ratio>=80
    add("CFO ≥ 80% of Net Profit (earnings quality)","Dhandho",cfo_np_ok, 5,
        f"{cfo_np_ratio:.0f}%" if not nn(cfo_np_ratio) else "N/A","≥ 80%",
        note="Cash-backed profits — high earnings quality" if cfo_np_ok else "Profits not converting to cash — accrual risk")

    # MoS using FCFF DCF
    mos = np.nan; intrinsic_cr = np.nan
    if current_price and not nn(current_price):
        avg_fcff = smean([v for v in fcff_v if not nn(v) and v>0])
        if nn(avg_fcff): avg_fcff = smean([v for v in fcf_v if not nn(v) and v>0])
        if not nn(avg_fcff) and not nn(l_sc) and l_sc>0:
            dr = max(wacc, 0.10)
            pv = sum(avg_fcff*(1.10**t)/(1+dr)**t for t in range(1,11))
            terminal = avg_fcff*(1.10**10)*12/(1+dr)**10
            intrinsic_cr = pv + terminal
            for face in [1,2,5,10]:
                shares = l_sc/face
                if shares>0.1:
                    mktcap = current_price*shares
                    if mktcap>0 and intrinsic_cr>0:
                        mos = (intrinsic_cr-mktcap)/intrinsic_cr*100
                    break
        mos_ok = not nn(mos) and mos>=30
        add("Margin of Safety ≥ 30% (FCFF DCF)","Dhandho",mos_ok, 8,
            f"MoS={mos:.1f}%" if not nn(mos) else "Cannot compute","≥ 30%",
            note="Significant discount to intrinsic value" if mos_ok else "Limited or negative margin of safety")

    # ════════════════════════════════════════════════════════════════════
    # SECTION E — DAMODARAN (advanced valuation)
    # ════════════════════════════════════════════════════════════════════

    # EV/EBITDA
    ev = np.nan
    if not nn(l_td) and not nn(l_cash) and current_price and not nn(l_sc):
        for face in [1,2,5,10]:
            shares = l_sc/face
            if shares>0.1:
                mktcap_ev = current_price*shares
                ev = mktcap_ev + l_td - l_cash
                break
    if nn(ev) and not nn(l_td) and not nn(l_np):
        # fallback: rough EV from PE * NP + Debt - Cash
        if not nn(l_pe) and not nn(l_np) and not nn(l_td):
            ev = l_pe*l_np + l_td - (l_cash if not nn(l_cash) else 0)
    ev_ebitda = sdiv(ev, l_ebitda) if not nn(ev) else np.nan
    ev_ok = not nn(ev_ebitda) and ev_ebitda<=12
    add("EV/EBITDA ≤ 12× (capital-structure neutral)","Damodaran",ev_ok, 6,
        f"{ev_ebitda:.1f}×" if not nn(ev_ebitda) else "N/A","≤ 12×",
        note=("Cheap: EV/EBITDA ≤ 8×" if not nn(ev_ebitda) and ev_ebitda<=8 else "Fair valuation on EBITDA basis") if ev_ok else "Expensive on EV/EBITDA basis")

    # PEG ratio
    eps_clean = [v for v in eps_v if not nn(v)]
    eps_cagr = np.nan
    if len(eps_clean)>=3 and eps_clean[0]>0 and eps_clean[-1]>0:
        eps_cagr = (eps_clean[-1]/eps_clean[0])**(1/(len(eps_clean)-1))-1
    peg = sdiv(l_pe, eps_cagr*100) if not nn(eps_cagr) and eps_cagr>0 else np.nan
    peg_ok = not nn(peg) and peg<=1.0
    add("PEG Ratio ≤ 1.0 (growth-adjusted PE)","Damodaran",peg_ok, 5,
        f"{peg:.2f}" if not nn(peg) else "N/A","≤ 1.0",
        note="Growth not yet priced in — undervalued relative to growth" if peg_ok else ("PEG 1–2: fair" if not nn(peg) and peg<=2 else "Growth fully/over-priced"))

    # P/B justified by ROE (Damodaran: fair P/B ≈ ROE / Cost of Equity)
    fair_pb = sdiv(smean([v for v in roe_v if not nn(v)]), cost_of_equity*100)
    pb_justified = not nn(fair_pb) and not nn(l_pb) and l_pb<=fair_pb*1.2
    add("P/B justified by ROE (P/B ≤ ROE/CoE)","Damodaran",pb_justified, 5,
        f"P/B={l_pb:.2f}× vs fair={fair_pb:.2f}×" if not nn(fair_pb) and not nn(l_pb) else "N/A",
        "P/B ≤ Fair P/B",
        note="Price supported by return on equity" if pb_justified else "P/B higher than ROE justifies — potential overvaluation")

    # ROIC Spread > 0 (value creation)
    spread_ok = not nn(roic_spread) and roic_spread>0
    add("ROIC Spread > 0% (ROCE > WACC — value creation)","Damodaran",spread_ok, 7,
        f"Spread={roic_spread:+.1f}pp (ROCE {l_roce:.1f}% vs WACC~{wacc_pct:.1f}%)" if not nn(roic_spread) else "N/A",
        "> 0pp",
        note=f"Creating value: ROCE exceeds WACC by {roic_spread:.1f}pp" if spread_ok else "Destroying value: ROCE < WACC — growth is dilutive")

    # EVA > 0
    eva_ok = not nn(eva) and eva>0
    add("EVA > 0 (Economic Value Added positive)","Damodaran",eva_ok, 6,
        f"₹{eva:.0f} Cr" if not nn(eva) else "N/A","> 0",
        note="Firm earning above its cost of capital" if eva_ok else "Economic profit negative — not covering cost of capital")

    # CFO / EBIT ≥ 1.0 (earnings quality — Damodaran)
    cfo_ebit = sdiv(l_cfo, l_op)
    cfo_ebit_ok = not nn(cfo_ebit) and cfo_ebit>=1.0
    add("CFO / EBIT ≥ 1.0 (Damodaran earnings quality)","Damodaran",cfo_ebit_ok, 5,
        f"{cfo_ebit:.2f}" if not nn(cfo_ebit) else "N/A","≥ 1.0",
        note="Operating cash fully backs reported earnings" if cfo_ebit_ok else "Cash conversion below EBIT — accrual-heavy reporting")

    # Working Capital as % of Revenue — should be stable or declining
    wc_pct_clean = [v for v in wc_pct_v if not nn(v)]
    sl_wc = slope(wc_pct_clean)
    wc_ok = not nn(sl_wc) and sl_wc<=0.5  # slight tolerance
    add("Working Capital / Revenue stable or declining","Damodaran",wc_ok, 4,
        f"{'Stable/Declining' if wc_ok else 'Increasing'} ({sl_wc:+.2f}pp/yr)" if not nn(sl_wc) else "N/A",
        "Stable or falling",
        note="Not consuming excess cash to fund growth" if wc_ok else "Rising WC intensity — cash trap risk")

    # Synthetic credit rating (Damodaran Ch.8)
    rating, spread_bps = synthetic_rating(ic)
    inv_grade = rating in ["AAA","AA","A","A−","BBB"]
    add(f"Synthetic credit: ≥ BBB (Damodaran IC→rating)","Damodaran",inv_grade, 4,
        f"{rating} (IC={ic:.1f}×)" if not nn(ic) and ic!=999 else (f"{rating} (Debt-free)" if ic==999 else "N/A"),
        "≥ BBB investment grade",
        note=f"Investment-grade equivalent ({rating})" if inv_grade else f"Sub-investment grade equivalent ({rating}) — elevated credit risk")

    # FCFF positive and growing
    fcff_clean = [v for v in fcff_v if not nn(v)]
    fcff_pos_count = sum(1 for v in fcff_clean if v>0)
    fcff_growing = len(fcff_clean)>=2 and slope(fcff_clean)>0
    fcff_ok = fcff_pos_count>0 and fcff_pos_count/max(len(fcff_clean),1)>=0.6 and fcff_growing
    add("FCFF positive & growing (Damodaran firm value)","Damodaran",fcff_ok, 5,
        f"{'Growing' if fcff_growing else 'Flat/Declining'}, {fcff_pos_count}/{len(fcff_clean)} yrs positive",
        "Positive & growing",
        note="Free cash flow to firm expanding — supports DCF value growth" if fcff_ok else "FCFF not consistently positive or growing")

    # ── VERDICT ──────────────────────────────────────────────────────
    pct = result["score"]/result["max_score"]*100 if result["max_score"]>0 else 0
    result["pct"] = pct
    if pct>=75:   result["verdict"]="STRONG BUY";  result["vcolor"]="#2a5c45"
    elif pct>=60: result["verdict"]="WATCHLIST";   result["vcolor"]="#d4850a"
    elif pct>=45: result["verdict"]="CAUTION";     result["vcolor"]="#d4850a"
    else:         result["verdict"]="AVOID";       result["vcolor"]="#c8401e"

    # ── derived data for charts ───────────────────────────────────────
    result["computed"] = dict(
        years=years, revenue=rev_v, net_profit=np_v, operating_profit=op_v,
        opm=opm_v, npm=npm_v, roce=roce_v, fcf=fcf_v, fcff=fcff_v,
        cfo=cfo_v, de=de_v, pe=pe_v, pb=pb_v, div=div_v, eps=eps_v,
        ebitda=ebitda_v, wc_pct=wc_pct_v, nopat=nopat_v, ce=ce_v,
    )
    result["derived"] = dict(
        wacc=wacc_pct, roic_spread=roic_spread, eva=eva,
        ev_ebitda=ev_ebitda, peg=peg, fair_pb=fair_pb,
        rating=rating, eps_cagr=eps_cagr, rev_cagr=rev_cagr,
        mos=mos, intrinsic_cr=intrinsic_cr,
        cost_of_equity=cost_of_equity*100,
    )
    return result

# ── CHART HELPERS ─────────────────────────────────────────────────────────────
def blayout(h=260):
    return dict(paper_bgcolor=PAPER,plot_bgcolor=PAPER,
        font=dict(family="DM Sans",color=INK,size=11),
        height=h, margin=dict(l=40,r=20,t=36,b=28),
        xaxis=dict(showgrid=False,color=MUTED,tickfont=dict(size=9)),
        yaxis=dict(gridcolor="#e8e3dc",color=MUTED,tickfont=dict(size=9)),
        showlegend=True,legend=dict(font=dict(size=9),orientation="h",y=1.15,x=0))

def bar_line(yrs,bars,lines,bn,ln,bc=A2,lc=A1):
    fig=make_subplots(specs=[[{"secondary_y":True}]])
    fig.add_trace(go.Bar(x=yrs,y=bars,name=bn,marker_color=bc,opacity=0.85),secondary_y=False)
    fig.add_trace(go.Scatter(x=yrs,y=lines,name=ln,line=dict(color=lc,width=2),
        mode="lines+markers",marker=dict(size=5)),secondary_y=True)
    fig.update_layout(**blayout())
    fig.update_yaxes(gridcolor="#e8e3dc",tickfont=dict(size=9),secondary_y=False)
    fig.update_yaxes(gridcolor="rgba(0,0,0,0)",tickfont=dict(size=9),secondary_y=True)
    return fig

def lines(yrs,series,h=260):
    cols=[A1,A2,BLUE,PURPLE,MUTED]
    fig=go.Figure()
    for i,(n,v) in enumerate(series.items()):
        fig.add_trace(go.Scatter(x=yrs,y=v,name=n,
            line=dict(color=cols[i%len(cols)],width=2),
            mode="lines+markers",marker=dict(size=5)))
    fig.update_layout(**blayout(h))
    return fig

def bars(yrs,vals,name,color=A2,h=260):
    cs=[A1 if not np.isnan(v) and v<0 else color for v in vals]
    fig=go.Figure(go.Bar(x=yrs,y=vals,name=name,marker_color=cs,opacity=0.85))
    fig.update_layout(**blayout(h))
    return fig

# ── MAIN APP ──────────────────────────────────────────────────────────────────
def main():
    st.markdown('<div class="main-header">Fundamental Analyser <span style="font-size:1rem;color:#7a7570">v2</span></div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Graham · Fisher · Siegel · Dhandho · Damodaran — 29 Rules</div>', unsafe_allow_html=True)

    # ── SIDEBAR ──────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### Upload & Settings")
        uploaded = st.file_uploader("Screener.in Export (CSV or Excel)", type=["csv","xlsx","xls"])
        st.markdown("---")
        st.markdown("##### Market Inputs")
        current_price = st.number_input("Current Market Price (₹)", min_value=0.0, value=0.0, step=1.0, format="%.2f")
        bond_yield    = st.number_input("10-yr Bond Yield (%)", min_value=0.0, value=7.0, step=0.1, format="%.1f")
        beta_input    = st.number_input("Beta (default 1.0)", min_value=0.1, value=1.0, step=0.05, format="%.2f")
        erp_input     = st.number_input("Equity Risk Premium (%)", min_value=1.0, value=5.5, step=0.5, format="%.1f")
        st.markdown("---")
        st.markdown("""<div style="font-size:0.72rem;color:#7a7570;line-height:1.7">
<b>Export from Screener.in:</b><br>
1. Open any company page<br>
2. Scroll to bottom<br>
3. Click <b>Export to Excel</b><br>
4. Upload the .xlsx file directly
</div>""", unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("""<div style="font-size:0.68rem;color:#7a7570;line-height:1.8">
🔵 <b>Graham</b> — Security Analysis<br>
🟢 <b>Fisher</b> — Common Stocks<br>
🟡 <b>Siegel</b> — Stocks for Long Run<br>
🔴 <b>Dhandho</b> — Pabrai<br>
🔷 <b>Damodaran</b> — Investment Valuation
</div>""", unsafe_allow_html=True)

    if not uploaded:
        # Landing
        st.markdown("---")
        c1,c2,c3,c4,c5 = st.columns(5)
        for col,val,lbl in zip([c1,c2,c3,c4,c5],
            ["29","5","FCFF","EVA","WACC"],
            ["Total Rules","Source Books","DCF Valuation","Economic Profit","Cost of Capital"]):
            with col:
                st.markdown(f'<div class="mcard"><div class="mval">{val}</div><div class="mlbl">{lbl}</div></div>',unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.info("👆 Upload a Screener.in CSV from the sidebar to begin analysis.")
        return

    # ── PARSE & ANALYSE ───────────────────────────────────────────────
    try:
        data = parse_screener_csv(uploaded)
    except Exception as e:
        st.error(f"Parse error: {e}"); st.stop()
    if not data["years"]:
        st.error("No year columns found. Check CSV format."); st.stop()

    # ── AUTO-DETECT PRICE FROM SCREENER META ────────────────────────
    # Screener.in xlsx stores Current Price in Data Sheet meta row
    meta_price = np.nan
    if "__meta__" in data.get("_parsed_meta", {}):
        pass
    # Try to get from data dict directly
    screener_price = data.get("screener_price", np.nan)
    if not np.isnan(screener_price) and current_price == 0:
        suggested_price = screener_price
    else:
        suggested_price = current_price if current_price > 0 else np.nan

    price_in = suggested_price if not np.isnan(suggested_price) else None
    result   = run_analysis(data, price_in, bond_yield, erp_input, beta_input)

    co   = data["company_name"]
    yrs  = data["years"]
    comp = result["computed"]
    derv = result["derived"]

    # ── HEADER STRIP ─────────────────────────────────────────────────
    st.markdown(f"## {co}")
    st.markdown(f'<span class="stag">{yrs[0]} – {yrs[-1]}</span> &nbsp; <span class="stag">{len(yrs)} years</span> &nbsp; <span class="stag">29 rules applied</span>', unsafe_allow_html=True)

    # ── PRICE INPUT STRIP — prominent on main page ────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    px1, px2, px3, px4 = st.columns([1.2, 1.2, 1.2, 3.4])

    with px1:
        # Show Screener's stored price as default if available
        default_price = float(screener_price) if not np.isnan(screener_price) else 0.0
        live_price = st.number_input(
            "📌 Current Market Price (₹)",
            min_value=0.0,
            value=default_price,
            step=0.5,
            format="%.2f",
            help="Enter the current stock price to unlock PE, EV/EBITDA, Margin of Safety calculations"
        )
        if live_price > 0:
            price_in = live_price
            # Re-run analysis with updated price
            result = run_analysis(data, price_in, bond_yield, erp_input, beta_input)
            comp = result["computed"]
            derv = result["derived"]

    with px2:
        live_bond = st.number_input(
            "📌 Bond Yield (%)",
            min_value=0.0,
            value=bond_yield,
            step=0.1,
            format="%.1f",
            help="10-year government bond yield — used for earnings yield comparison"
        )
        if live_bond != bond_yield and live_price > 0:
            result = run_analysis(data, price_in, live_bond, erp_input, beta_input)
            comp = result["computed"]
            derv = result["derived"]
        elif live_bond != bond_yield:
            result = run_analysis(data, price_in, live_bond, erp_input, beta_input)
            comp = result["computed"]
            derv = result["derived"]

    with px3:
        if not np.isnan(screener_price) and default_price > 0:
            st.markdown(f"""<div class="mcard" style="margin-top:1.8rem">
  <div style="font-family:'DM Mono',monospace;font-size:0.6rem;color:#7a7570;text-transform:uppercase;letter-spacing:0.1em">Screener Price</div>
  <div style="font-family:'Instrument Serif',serif;font-size:1.4rem;font-style:italic">₹{screener_price:.2f}</div>
  <div style="font-size:0.65rem;color:#7a7570;margin-top:0.1rem">Auto-filled from file</div>
</div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""<div class="mcard" style="margin-top:1.8rem;border-left:3px solid #d4850a">
  <div style="font-size:0.75rem;color:#d4850a">⚠️ Enter price to unlock<br>PE · EV/EBITDA · Margin of Safety</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── KEY METRICS ROW ───────────────────────────────────────────────
    m1,m2,m3,m4,m5,m6 = st.columns(6)
    def mcard(col,val,lbl,color=INK):
        with col:
            st.markdown(f'<div class="mcard"><div class="mval" style="color:{color}">{val}</div><div class="mlbl">{lbl}</div></div>',unsafe_allow_html=True)

    l_pe  = latest(comp["pe"]);  l_pb = latest(comp["pb"])
    l_roce= latest(comp["roce"]); l_de  = latest(comp["de"])
    mcard(m1, f"{result['pct']:.0f}%", "Overall Score", result["vcolor"])
    mcard(m2, f"{l_pe:.1f}×"   if not np.isnan(l_pe)   else "N/A", "P/E Ratio")
    mcard(m3, f"{l_roce:.1f}%" if not np.isnan(l_roce) else "N/A", "Latest ROCE")
    mcard(m4, f"{derv['wacc']:.1f}%", "Est. WACC")
    mcard(m5, f"{derv['roic_spread']:+.1f}pp" if not np.isnan(derv['roic_spread']) else "N/A",
          "ROIC Spread", "#2a5c45" if not np.isnan(derv['roic_spread']) and derv['roic_spread']>0 else "#c8401e")
    mcard(m6, derv.get("rating","N/A"), "Synthetic Rating",
          "#2a5c45" if derv.get("rating","N/A") in ["AAA","AA","A","A−","BBB"] else "#c8401e")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── TABS ─────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs(["📋 Score & Rules", "📈 Performance Charts", "🔬 Damodaran Analysis", "🗃️ Raw Data"])

    # ════════════════════════════════════════════════════════════════
    # TAB 1 — SCORE & RULES
    # ════════════════════════════════════════════════════════════════
    with tab1:
        col_gauge, col_rules, col_flags = st.columns([1, 2.2, 1.8])

        with col_gauge:
            fig_g = go.Figure(go.Indicator(
                mode="gauge+number", value=result["pct"],
                number={"suffix":"%","font":{"family":"Instrument Serif","size":34,"color":INK}},
                gauge={"axis":{"range":[0,100],"tickfont":{"size":8}},
                    "bar":{"color":result["vcolor"],"thickness":0.28},
                    "bgcolor":PAPER,"bordercolor":MUTED,
                    "steps":[{"range":[0,45],"color":"#fdf0ee"},
                              {"range":[45,60],"color":"#fef6e7"},
                              {"range":[60,75],"color":"#fef6e7"},
                              {"range":[75,100],"color":"#edf5f1"}]},
                title={"text":"SCORE","font":{"family":"DM Mono","size":10,"color":MUTED}}))
            fig_g.update_layout(paper_bgcolor=PAPER,plot_bgcolor=PAPER,
                height=220,margin=dict(l=20,r=20,t=28,b=8),font=dict(family="DM Sans"))
            st.plotly_chart(fig_g, use_container_width=True)
            st.markdown(f"""<div style="text-align:center;margin-top:-0.8rem">
  <span style="font-family:'Instrument Serif',serif;font-size:1.3rem;font-style:italic;color:{result['vcolor']}">{result['verdict']}</span><br>
  <span style="font-family:'DM Mono',monospace;font-size:0.68rem;color:{MUTED}">{result['score']:.0f} / {result['max_score']:.0f} pts</span>
</div>""", unsafe_allow_html=True)

            # Book breakdown bars
            st.markdown("<br>", unsafe_allow_html=True)
            book_sc = {}; book_mx = {}
            for r in result["rules"]:
                s = r["source"]
                book_sc[s] = book_sc.get(s,0) + r["pts"]
                book_mx[s] = book_mx.get(s,0) + r["max_pts"]
            for book in ["Graham","Fisher","Siegel","Dhandho","Damodaran"]:
                if book not in book_mx: continue
                got=book_sc.get(book,0); mx=book_mx[book]
                p=got/mx*100 if mx>0 else 0
                bc="#2a5c45" if p>=70 else ("#d4850a" if p>=40 else "#c8401e")
                st.markdown(f"""<div style="margin-bottom:0.5rem">
  <div style="display:flex;justify-content:space-between;font-family:'DM Mono',monospace;font-size:0.62rem;color:{MUTED}">
    <span>{book}</span><span style="color:{bc}">{p:.0f}%</span></div>
  <div style="background:#e8e3dc;border-radius:2px;height:5px;margin-top:2px">
    <div style="background:{bc};width:{int(p)}%;height:5px;border-radius:2px"></div></div>
</div>""", unsafe_allow_html=True)

        with col_rules:
            st.markdown('<div class="stag">All 29 Rules — click ▶ to expand explanation</div>', unsafe_allow_html=True)

            def find_explanation(rule_name):
                for key, val in RULE_EXPLANATIONS.items():
                    if rule_name.lower().startswith(key.lower()) or key.lower() in rule_name.lower():
                        return val
                return None

            for book in ["Graham","Fisher","Siegel","Dhandho","Damodaran"]:
                book_rules = [r for r in result["rules"] if r["source"]==book]
                if not book_rules: continue
                c = SOURCE_COLORS[book]
                st.markdown(f'<div style="font-family:DM Mono,monospace;font-size:0.6rem;color:{c};letter-spacing:0.1em;text-transform:uppercase;margin-top:0.8rem;margin-bottom:0.2rem;border-left:2px solid {c};padding-left:6px">{book}</div>', unsafe_allow_html=True)
                for r in book_rules:
                    icon  = "\u2705" if r["passed"] else "\u274c"
                    expl  = find_explanation(r["name"])
                    st.markdown(f"""<div class="rule-row">
  <span class="rule-icon">{icon}</span>
  <span class="rule-text">{r['name']}</span>
  <span class="rule-val">{r['actual']}</span>
</div>""", unsafe_allow_html=True)
                    if expl:
                        abbrev, formula, plain = expl
                        border_c = "#2a5c45" if r["passed"] else "#c8401e"
                        with st.expander(f"\U0001f4d6  {abbrev}", expanded=False):
                            st.markdown(f"""<div style="font-size:0.82rem;line-height:1.65">
  <div style="margin-bottom:0.6rem">
    <span style="font-family:DM Mono,monospace;font-size:0.6rem;letter-spacing:0.1em;text-transform:uppercase;color:#7a7570">FORMULA</span><br>
    <code style="background:#f0ede6;padding:3px 7px;border-radius:3px;font-size:0.79rem">{formula}</code>
  </div>
  <div style="border-left:3px solid {border_c};padding:0.4rem 0 0.4rem 0.8rem;color:#2a2825;background:{"#edf5f1" if r["passed"] else "#fdf0ee"};border-radius:0 3px 3px 0">
    {plain}
  </div>
</div>""", unsafe_allow_html=True)


        with col_flags:
            positives = result["positives"][:8]
            flags     = result["flags"][:8]
            if positives:
                st.markdown('<div class="stag">Strengths</div>', unsafe_allow_html=True)
                for p in positives:
                    src,msg = p.split("] ",1) if "] " in p else ("",p)
                    src=src.replace("[","")
                    st.markdown(f'<div class="flag-green"><b style="font-size:0.67rem;color:#2a5c45">{src}</b><br>{msg}</div>',unsafe_allow_html=True)
            if flags:
                st.markdown('<div class="stag" style="margin-top:0.8rem">Red Flags</div>', unsafe_allow_html=True)
                for f in flags:
                    src,msg = f.split("] ",1) if "] " in f else ("",f)
                    src=src.replace("[","")
                    st.markdown(f'<div class="flag-red"><b style="font-size:0.67rem;color:#c8401e">{src}</b><br>{msg}</div>',unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════
    # TAB 2 — PERFORMANCE CHARTS
    # ════════════════════════════════════════════════════════════════
    with tab2:
        r1c1, r1c2 = st.columns(2)
        with r1c1:
            st.markdown('<div class="stag">Revenue & Net Profit (Cr)</div>', unsafe_allow_html=True)
            st.plotly_chart(bar_line(yrs,comp["revenue"],comp["net_profit"],"Revenue","Net Profit"),use_container_width=True)
        with r1c2:
            st.markdown('<div class="stag">Operating & Net Margin %</div>', unsafe_allow_html=True)
            st.plotly_chart(lines(yrs,{"OPM %":comp["opm"],"NPM %":comp["npm"]}),use_container_width=True)

        r2c1, r2c2 = st.columns(2)
        with r2c1:
            st.markdown('<div class="stag">ROCE % — Moat Indicator</div>', unsafe_allow_html=True)
            fig = bars(yrs,comp["roce"],"ROCE %",A2)
            fig.add_hline(y=15,line_dash="dash",line_color=MUTED,
                annotation_text="Moat threshold (15%)",annotation_position="top right",
                annotation_font=dict(size=9,color=MUTED))
            fig.add_hline(y=20,line_dash="dot",line_color=A2,
                annotation_text="Strong moat (20%)",annotation_position="bottom right",
                annotation_font=dict(size=9,color=A2))
            st.plotly_chart(fig,use_container_width=True)
        with r2c2:
            st.markdown('<div class="stag">Free Cash Flow (Cr)</div>', unsafe_allow_html=True)
            st.plotly_chart(bars(yrs,comp["fcf"],"FCF",BLUE),use_container_width=True)

        r3c1, r3c2 = st.columns(2)
        with r3c1:
            st.markdown('<div class="stag">Debt/Equity Ratio</div>', unsafe_allow_html=True)
            fig = bars(yrs,comp["de"],"D/E",A1)
            fig.add_hline(y=1.0,line_dash="dash",line_color=MUTED,
                annotation_text="Graham safe limit (1.0)",annotation_font=dict(size=9,color=MUTED))
            st.plotly_chart(fig,use_container_width=True)
        with r3c2:
            st.markdown('<div class="stag">PE Ratio History</div>', unsafe_allow_html=True)
            fig = bars(yrs,comp["pe"],"PE",PURPLE)
            fig.add_hline(y=20,line_dash="dash",line_color=A1,
                annotation_text="Graham cap (20×)",annotation_font=dict(size=9,color=A1))
            fig.add_hline(y=15,line_dash="dot",line_color=A2,
                annotation_text="Fair (15×)",annotation_font=dict(size=9,color=A2))
            st.plotly_chart(fig,use_container_width=True)

    # ════════════════════════════════════════════════════════════════
    # TAB 3 — DAMODARAN ANALYSIS
    # ════════════════════════════════════════════════════════════════
    with tab3:
        st.markdown("#### Damodaran Valuation Framework")

        # Top metric strip
        dm1,dm2,dm3,dm4,dm5 = st.columns(5)
        def dcard(col,val,lbl,note="",color=INK):
            with col:
                st.markdown(f'<div class="mcard"><div class="mval" style="color:{color}">{val}</div>'
                    f'<div class="mlbl">{lbl}</div>'
                    f'{"<div style=\"font-size:0.68rem;color:#7a7570;margin-top:0.3rem\">"+note+"</div>" if note else ""}'
                    f'</div>',unsafe_allow_html=True)

        ev_eb = derv["ev_ebitda"]; peg_v = derv["peg"]
        spread = derv["roic_spread"]; eva_v = derv["eva"]
        dcard(dm1, f"{ev_eb:.1f}×"    if not np.isnan(ev_eb)  else "N/A", "EV/EBITDA",
              "≤8× cheap, ≤12× fair",
              "#2a5c45" if not np.isnan(ev_eb) and ev_eb<=12 else "#c8401e")
        dcard(dm2, f"{peg_v:.2f}"     if not np.isnan(peg_v)  else "N/A", "PEG Ratio",
              "< 1.0 = undervalued",
              "#2a5c45" if not np.isnan(peg_v) and peg_v<=1 else "#c8401e")
        dcard(dm3, f"{spread:+.1f}pp" if not np.isnan(spread) else "N/A", "ROIC Spread",
              f"WACC={derv['wacc']:.1f}%",
              "#2a5c45" if not np.isnan(spread) and spread>0 else "#c8401e")
        dcard(dm4, f"₹{eva_v:.0f}Cr"  if not np.isnan(eva_v)  else "N/A", "EVA",
              "Econ. Value Added",
              "#2a5c45" if not np.isnan(eva_v) and eva_v>0 else "#c8401e")
        dcard(dm5, derv.get("rating","N/A"), "Synthetic Rating",
              f"IC-based (Damodaran)",
              "#2a5c45" if derv.get("rating","N/A") in ["AAA","AA","A","A−","BBB"] else "#c8401e")

        st.markdown("<br>", unsafe_allow_html=True)

        # Charts row 1
        dc1, dc2 = st.columns(2)
        with dc1:
            st.markdown('<div class="stag">NOPAT & Capital Employed (Cr)</div>', unsafe_allow_html=True)
            st.plotly_chart(bar_line(yrs,comp["ce"],comp["nopat"],"Capital Employed","NOPAT",bc=BLUE,lc=A2),use_container_width=True)
        with dc2:
            st.markdown('<div class="stag">FCFF — Free Cash Flow to Firm (Cr)</div>', unsafe_allow_html=True)
            st.plotly_chart(bars(yrs,comp["fcff"],"FCFF",A2),use_container_width=True)

        # Charts row 2
        dc3, dc4 = st.columns(2)
        with dc3:
            st.markdown('<div class="stag">EBITDA (Cr) — EV/EBITDA base</div>', unsafe_allow_html=True)
            st.plotly_chart(bars(yrs,comp["ebitda"],"EBITDA",PURPLE),use_container_width=True)
        with dc4:
            st.markdown('<div class="stag">Working Capital % of Revenue</div>', unsafe_allow_html=True)
            fig = lines(yrs,{"WC % Rev":comp["wc_pct"]})
            st.plotly_chart(fig,use_container_width=True)

        # DCF Intrinsic Value summary box
        st.markdown("---")
        st.markdown("#### Intrinsic Value Estimate (Damodaran FCFF DCF)")
        iv1,iv2,iv3,iv4 = st.columns(4)
        mos = derv.get("mos",np.nan); ic_val = derv.get("intrinsic_cr",np.nan)
        dcard(iv1, f"{derv['cost_of_equity']:.1f}%", "Cost of Equity", f"Rf={bond_yield}% + β{beta_input}×ERP{erp_input}%")
        dcard(iv2, f"{derv['wacc']:.1f}%", "Est. WACC", "Approx. from IC & D/E")
        dcard(iv3, f"₹{ic_val:.0f} Cr"  if not np.isnan(ic_val) else "N/A", "DCF Intrinsic (Firm)", "10-yr FCFF @ 10% growth")
        dcard(iv4, f"{mos:.1f}%"         if not np.isnan(mos)   else "N/A", "Margin of Safety",
              "vs current mkt price",
              "#2a5c45" if not np.isnan(mos) and mos>=30 else "#c8401e")

        if np.isnan(mos):
            st.info("ℹ️ Enter a **Current Market Price** in the sidebar to compute Margin of Safety.")

        # Synthetic rating table
        st.markdown("---")
        st.markdown("#### Synthetic Credit Rating (Damodaran IC → Rating Map)")
        rat_data = {
            "Interest Coverage": [">8.5×","6.5–8.5×","4.5–6.5×","3.5–4.5×","2.5–3.5×","1.5–2.5×","<1.5×"],
            "Synthetic Rating":  ["AA","A","A−","BBB","BB","B","CCC"],
            "Default Spread":    ["0.70%","1.07%","1.21%","1.47%","2.21%","3.61%","8.51%"],
        }
        rt = derv.get("rating","N/A")
        l_ic_disp = latest(av(data["interest"], yrs))
        ic_disp = latest([latest(av(data["operating_profit"],yrs))/l_ic_disp]) if not np.isnan(l_ic_disp) and l_ic_disp>0 else np.nan
        st.dataframe(pd.DataFrame(rat_data).set_index("Interest Coverage"),use_container_width=True)
        st.markdown(f"**This company's estimated rating: `{rt}`** — based on latest interest coverage ratio.")

    # ════════════════════════════════════════════════════════════════
    # TAB 4 — RAW DATA
    # ════════════════════════════════════════════════════════════════
    with tab4:
        st.markdown('<div class="stag">Parsed Financial Data</div>', unsafe_allow_html=True)
        metrics = {
            "Revenue (Cr)":        comp["revenue"],
            "Operating Profit (Cr)":comp["operating_profit"],
            "Net Profit (Cr)":     comp["net_profit"],
            "EBITDA (Cr)":         comp["ebitda"],
            "EPS (₹)":             comp["eps"],
            "OPM %":               comp["opm"],
            "NPM %":               comp["npm"],
            "ROCE %":              comp["roce"],
            "D/E Ratio":           comp["de"],
            "PE Ratio":            comp["pe"],
            "P/B Ratio":           comp["pb"],
            "CFO (Cr)":            comp["cfo"],
            "FCF (Cr)":            comp["fcf"],
            "FCFF (Cr)":           comp["fcff"],
            "NOPAT (Cr)":          comp["nopat"],
            "Cap. Employed (Cr)":  comp["ce"],
            "WC % Revenue":        comp["wc_pct"],
            "Dividend (Cr)":       comp["div"],
        }
        rows = {}
        for metric,vals in metrics.items():
            rows[metric] = {y: (round(v,2) if not np.isnan(v) else "—") for y,v in zip(yrs,vals)}
        st.dataframe(pd.DataFrame(rows).T, use_container_width=True)

        st.markdown("---")
        st.markdown('<div class="stag">Derived Valuation Metrics</div>', unsafe_allow_html=True)
        dv = derv
        derived_display = {
            "Metric": ["WACC (Est.%)", "Cost of Equity %", "ROIC Spread (pp)", "EVA (Cr)",
                       "EV/EBITDA", "PEG Ratio", "Fair P/B (ROE/CoE)", "Revenue CAGR %",
                       "EPS CAGR %", "Margin of Safety %", "Synthetic Credit Rating"],
            "Value": [
                f"{dv['wacc']:.2f}%", f"{dv['cost_of_equity']:.2f}%",
                f"{dv['roic_spread']:+.2f}pp" if not np.isnan(dv['roic_spread']) else "N/A",
                f"₹{dv['eva']:.0f} Cr" if not np.isnan(dv['eva']) else "N/A",
                f"{dv['ev_ebitda']:.2f}×" if not np.isnan(dv['ev_ebitda']) else "N/A",
                f"{dv['peg']:.2f}" if not np.isnan(dv['peg']) else "N/A",
                f"{dv['fair_pb']:.2f}×" if not np.isnan(dv['fair_pb']) else "N/A",
                f"{dv['rev_cagr']*100:.1f}%" if not np.isnan(dv['rev_cagr']) else "N/A",
                f"{dv['eps_cagr']*100:.1f}%" if not np.isnan(dv['eps_cagr']) else "N/A",
                f"{dv['mos']:.1f}%" if not np.isnan(dv['mos']) else "Enter price in sidebar",
                dv.get("rating","N/A"),
            ]
        }
        st.dataframe(pd.DataFrame(derived_display).set_index("Metric"), use_container_width=True)

    # ── FOOTER ───────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("""<div style="font-family:'DM Mono',monospace;font-size:0.62rem;color:#7a7570;text-align:center;padding:0.8rem 0">
Security Analysis (Graham & Dodd) · Common Stocks & Uncommon Profits (Fisher) · Stocks for the Long Run (Siegel) · 
The Dhandho Investor (Pabrai) · Investment Valuation (Damodaran 4th Ed.) · Screening tool only — not financial advice.
</div>""", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
